"""
Performance Analysis Tool for Unreal Engine CSV Exports
Robust CSV → Excel Report Generator

Author: Marvin Schubert
Version: 1.0.0
Date: September 2025

INSTALLATION REQUIREMENTS:
    pip install pandas numpy openpyxl

USAGE:
    1. Create a folder named 'messungen' in the script directory
    2. Place your EXP_*.csv files (Unreal Engine profiler exports) into this folder
    3. Run the script: python messung_auswertung.py
    4. Get Excel report with p95/mean values grouped by scene/variant
    
FEATURES:
    - Handles variable column counts in UE CSV exports
    - Calculates frame consistency metrics (p95 percentiles)
    - Automatic scene/variant/run detection from filenames
    - German-formatted Excel output with proper decimal separators
    
INPUT FORMAT:
    Expected CSV filename pattern: EXP_[SceneNumber]_[A|B]_Messung_[RunNumber].csv
    Example: EXP_1_A_Messung_1.csv, EXP_2_B_Messung_3.csv
"""
import re, csv, datetime
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

# Column mapping for Unreal Engine CSV exports
# Maps logical metrics to possible column name variations
ALIASES = {
    "FrameTime": ["FrameTime (ms)", "Frame Time (ms)", "FrameTime", "Frame Time", "Frame (ms)", "FrameTimeMs", "Frame", "FrameMs"],
    "GPUTime": ["GPU (ms)", "GPUTime (ms)", "GPUTime", "GPU Time (ms)", "GPU"],
    "DrawCalls": ["Draw Calls", "RHI Draw Calls", "DrawCalls", "DrawCallCount", "DrawPrimitive", "RHI/DrawCalls"],
    "Primitives": ["Primitives", "RHI Primitives", "Visible Primitives", "VisiblePrimitives", "RHI/PrimitivesDrawn", "PrimitivesDrawn"],
    "VRAM": ["RHI GPU Memory (MB)", "GPU Memory Used (MB)", "RHI GPU Memory", "GPUMemoryMB", "GPU Memory (MB)", "GPUMem/LocalUsedMB", "LocalUsedMB"],
    "ShaderMem": ["Shader Mem (MB)", "Shader Mem", "Shader Memory (MB)", "Shader Memory", "ShaderMemMB", "ShaderMemoryMB", "ShaderMemory", "Shaders/ShaderMemoryMB"],
}

# --------------------------------------------------
# Numeric string normalization helper
# Rules:
#  - Remove NBSP (\u00A0) and regular spaces used as thousand separators
#  - If both '.' and ',' exist and the last separator is ',', treat pattern like 12.345,678 -> 12345.678
#  - If only one comma (e.g. 3,141) and no dot, treat comma as decimal -> 3.141
#  - Leave other strings untouched; conversion to numeric happens later with coercion
def _normalize_numeric_cell(val):
    if not isinstance(val, str):
        return val
    s = val.replace('\u00A0', ' ').strip()
    if not s:
        return val
    # Remove spaces (thousand separators) now, but keep sign
    # Keep a leading minus or plus if present
    sign = ''
    if s[0] in '+-':
        sign, s = s[0], s[1:]
    s = s.replace(' ', '')

    if any(ch.isdigit() for ch in s):
        if '.' in s and ',' in s:
            # Decide which is decimal: assume European style if last separator is ','
            if s.rfind(',') > s.rfind('.'):
                # Remove dots (thousand) then replace comma with dot (decimal)
                s = s.replace('.', '')
                s = s.replace(',', '.')
            else:
                # Pattern like 12,345.678 -> remove commas
                s = s.replace(',', '')
        elif ',' in s and '.' not in s:
            # Single comma acts as decimal (common German style)
            if s.count(',') == 1:
                s = s.replace(',', '.')
            else:
                # Multiple commas -> remove all but keep last as decimal if plausible
                parts = s.split(',')
                if parts[-1].isdigit():
                    decimal = parts.pop()
                    s = ''.join(parts).replace('.', '') + '.' + decimal
                else:
                    s = s.replace(',', '')
        # Remove stray thousands separators like apostrophes
        s = s.replace("'", '')
        # Reattach sign
        s = sign + s
    return s

def _normalize(s: str) -> str:
    """Normalize string for column matching - remove special chars and lowercase"""
    return re.sub(r"[^a-z0-9]+", "", s.strip().lower())

def find_column(df: pd.DataFrame, candidates: list) -> str:
    """Find column in DataFrame using flexible matching with candidate names"""
    # Exact normalized matches first
    norm_cols = {_normalize(c): c for c in df.columns}
    for cand in candidates:
        n = _normalize(cand)
        if n in norm_cols:
            return norm_cols[n]
    
    # Partial matches as fallback
    for c in df.columns:
        if any(_normalize(cand) in _normalize(c) for cand in candidates):
            return c
    return None

def detect_header_start(path: Path) -> int:
    """Find the line where CSV header starts by looking for Frame-related keywords"""
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()
    for i, line in enumerate(lines[:200]):  # Search only first 200 lines
        if (",Frame" in line) or ("Frame," in line) or ("FrameTime" in line) or ("Frame Time" in line):
            return i
    return 0

def load_csv(path: Path) -> pd.DataFrame:
    """
    Robust CSV loading that handles variable column counts.
    
    UE CSV exports often have inconsistent column counts which causes pandas
    to skip "malformed" lines. This function handles this by:
    1. Manual line-by-line parsing
    2. Normalizing row lengths to match header
    3. Fallback to standard pandas methods if needed
    
    Args:
        path: Path to CSV file
        
    Returns:
        pandas DataFrame with loaded data
    """
    header_idx = detect_header_start(path)
    
    # Primary strategy: Robust line-by-line parsing for variable column counts
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()[header_idx:]
        
        if not lines:
            raise ValueError("No data lines found")
        
        # Auto-detect separator from header line
        header_line = lines[0].strip()
        sep = "," if header_line.count(",") > header_line.count(";") else ";"
        
        # Parse header to determine expected column count
        header_cols = [col.strip('"').strip() for col in header_line.split(sep)]
        max_cols = len(header_cols)
        
        # Parse data lines with variable column handling
        data_rows = []
        skipped_lines = 0
        
        for i, line in enumerate(lines[1:], start=2):
            line = line.strip()
            if not line:  # Skip empty lines
                continue
                
            try:
                # Split and clean row data
                row_data = [val.strip('"').strip() for val in line.split(sep)]
                
                # Normalize row length to match header
                if len(row_data) < max_cols:
                    # Extend with None for missing columns
                    row_data.extend([None] * (max_cols - len(row_data)))
                elif len(row_data) > max_cols:
                    # Truncate extra columns
                    row_data = row_data[:max_cols]
                
                data_rows.append(row_data)
                
            except Exception:
                skipped_lines += 1
                continue
        
        if not data_rows:
            raise ValueError("No valid data lines found")
        
        # Create and clean DataFrame
        df = pd.DataFrame(data_rows, columns=header_cols)
        df = df.dropna(axis=1, how="all")  # Remove completely empty columns
        df = df.replace("", None)  # Replace empty strings with None

        # ---------------- PATCH 1: Numeric string normalization + coercion ----------------
        obj_cols = [c for c in df.columns if df[c].dtype == 'object']
        for c in obj_cols:
            df[c] = df[c].map(_normalize_numeric_cell)
            tmp = pd.to_numeric(df[c], errors='coerce')
            if tmp.notna().any():
                df[c] = tmp
        # Drop columns that are now fully NaN
        df = df.dropna(axis=1, how='all')
        
        total_lines = len(lines) - 1  # -1 for header
        loaded_lines = len(data_rows)
        
        print(f"    Robustly loaded: {loaded_lines}/{total_lines} lines (Separator: '{sep}', {skipped_lines} skipped)")
        
        return df
        
    except Exception as e:
        print(f"    Robust method failed: {e}")
    
    # Fallback: Standard pandas methods with different configurations
    for sep in [",", ";", "\t"]:
        for engine in ["python", "c"]:
            try:
                df = pd.read_csv(path, skiprows=header_idx, sep=sep, engine=engine,
                               quoting=csv.QUOTE_NONE, escapechar="\\", 
                               on_bad_lines="skip", encoding='utf-8', 
                               dtype=str, na_values=[''])
                if len(df.columns) > 1 and len(df) > 0:
                    print(f"    Fallback: {len(df)} lines with {engine}-engine and '{sep}'-separator")
                    return df.dropna(axis=1, how="all")
            except Exception:
                continue
    
    raise RuntimeError(f"CSV {path.name} could not be loaded")

def p95(series: pd.Series) -> float:
    """Calculate 95th percentile - critical performance metric for frame consistency"""
    series = pd.to_numeric(series, errors="coerce").dropna()
    return float(np.percentile(series.to_numpy(), 95)) if not series.empty else float("nan")

def mean(series: pd.Series) -> float:
    """Calculate arithmetic mean with robust error handling"""
    series = pd.to_numeric(series, errors="coerce").dropna()
    return float(series.mean()) if not series.empty else float("nan")

def parse_filename(path: Path) -> tuple:
    """
    Extract scene, variant (A/B), run from filename like EXP_1_A_Messung_1.csv
    
    Returns:
        tuple: (scene, variant, run) as strings
    """
    m = re.search(r"EXP[_-]?(?P<scene>\d+).*?(?P<variant>[AB]).*?(?:Messung|Run|Lauf)[_-]?(?P<run>\d+)?", path.name, re.IGNORECASE)
    scene = m.group("scene") if m else "1"
    variant = m.group("variant").upper() if m and m.group("variant") else "?"
    run = m.group("run") if m and m.group("run") else "1"
    return scene, variant, run

def collect_metrics(df: pd.DataFrame) -> dict:
    """
    Extract all performance metrics from DataFrame.
    
    Calculates means and p95 values for frame timing and GPU metrics.
    Uses flexible column matching to handle varying UE export formats.
    
    Args:
        df: DataFrame with performance data
        
    Returns:
        dict: Computed metrics with debug info about found columns
    """
    # Define metrics to calculate: (aggregation_function, column_alias_key)
    metrics = {
        "FrameTime_mean": (mean, "FrameTime"),
        "FrameTime_p95": (p95, "FrameTime"),
        "GPUTime_mean": (mean, "GPUTime"),
        "GPUTime_p95": (p95, "GPUTime"),
        "DrawCalls_mean": (mean, "DrawCalls"),
        "Primitives_mean": (mean, "Primitives"),
        "VRAM_mean": (mean, "VRAM"),
        "ShaderMem_mean": (mean, "ShaderMem"),
    }
    
    # Initialize results with frame count
    values = {"Frames": int(len(df))}
    found_cols = {}  # Track which columns were found for debugging
    
    # Calculate each metric
    for metric_key, (agg_fn, alias_key) in metrics.items():
        col = find_column(df, ALIASES.get(alias_key, []))
        found_cols[metric_key] = col or "NOT_FOUND"
        values[metric_key] = agg_fn(df[col]) if col else float("nan")
    
    values["_debug_cols"] = found_cols
    return values

def format_number(label: str, value: float) -> str:
    """PATCH 4: Einheitliche Formatierung (DE)
    - Integer-Metriken ohne Nachkommastellen
    - Tausender = NBSP
    - Dezimal = Komma (3 Nachkommastellen)
    """
    if not np.isfinite(value):
        return ''
    NBSP = '\u00A0'

    def group_with_nbsp(n: int) -> str:
        s = str(abs(int(n)))
        parts = []
        while s:
            parts.insert(0, s[-3:])
            s = s[:-3]
        return ('-' if n < 0 else '') + NBSP.join(parts)

    def format_decimal(v: float, decimals: int = 3) -> str:
        rounded = round(float(v), decimals)
        int_part, _, dec_part = f"{rounded:.{decimals}f}".partition('.')
        int_fmt = group_with_nbsp(int(int_part))
        return f"{int_fmt},{dec_part}" if decimals > 0 else int_fmt

    integer_labels = {'N', 'Draw Calls Ø [#]', 'Primitives Ø [#]'}
    if label in integer_labels:
        return group_with_nbsp(int(round(value)))
    return format_decimal(value, decimals=3)

def main():
    """
    Main processing function:
    1. Scan for CSV files in 'messungen' directory
    2. Load and process each file with robust CSV handling
    3. Extract performance metrics (means, p95 values)
    4. Generate Excel report grouped by scene/variant
    """
    # Check for input directory
    messungen_dir = Path("messungen")
    if not messungen_dir.exists():
        print("Directory 'messungen' not found!")
        return
    
    # Find CSV files with expected naming pattern
    csv_files = list(messungen_dir.glob("EXP_*.csv"))
    if not csv_files:
        print("No EXP_*.csv files found in messungen directory!")
        return
    
    print(f"Processing {len(csv_files)} CSV files...")
    
    # Process each CSV file
    runs = []
    for path in csv_files:
        scene, variant, run = parse_filename(path)
        try:
            df = load_csv(path)
            metrics = collect_metrics(df)
            metrics.update({"scene": scene, "variant": variant, "run": run, "file": path.name})
            runs.append(metrics)
            print(f"  ✓ {path.name} -> Scene {scene}, Variant {variant}, Run {run}")
            
            # Show column detection results for troubleshooting
            debug_cols = metrics.get("_debug_cols", {})
            missing = [k for k, v in debug_cols.items() if v == "NOT_FOUND"]
            if missing:
                print(f"    ⚠ Missing columns: {', '.join(missing)}")
            else:
                print(f"    ✓ All metrics found")
        except Exception as e:
            print(f"  ✗ {path.name}: {e}")
    
    if not runs:
        print("No processable files found!")
        return
    
    # Generate Excel report
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Prepare data for Excel export
        df_runs = pd.DataFrame(runs)
        df_runs['run_ord'] = df_runs['run'].astype(str).str.extract(r'(\d+)').astype(int)
        # ---------------- PATCH 3: Robust FPS derivation ----------------
        ft = df_runs['FrameTime_mean']
        df_runs['FPS_mean'] = np.where((ft > 0) & np.isfinite(ft), 1000.0 / ft, np.nan)

        # Define metrics to include in Excel report
        metric_rows = [
            ('N', 'Frames'),
            ('Frametime Ø [ms]', 'FrameTime_mean'),
            ('Frametime p95 [ms]', 'FrameTime_p95'),
            ('FPS Ø [#]', 'FPS_mean'),
            ('GPU Zeit Ø [ms]', 'GPUTime_mean'),
            ('GPU Zeit p95 [ms]', 'GPUTime_p95'),
            ('Draw Calls Ø [#]', 'DrawCalls_mean'),
            ('Primitives Ø [#]', 'Primitives_mean'),
            ('Local VRAM [MB]', 'VRAM_mean'),
            ('Shader Mem [MB]', 'ShaderMem_mean'),
        ]
        
        # Create worksheet for each scene/variant combination
        for (scene, variant), grp in df_runs.groupby(['scene', 'variant']):
            ws = wb.create_sheet(title=f"Scene{scene}_{variant}")
            grp_sorted = grp.sort_values(['run_ord'])
            runs_list = list(grp_sorted['run'])
            
            # Create header row
            ws.cell(row=1, column=1, value=f"Measurements – Variant {variant}").font = Font(bold=True)
            for ci, rname in enumerate(runs_list, start=2):
                c = ws.cell(row=1, column=ci, value=f"Run {rname}")
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
            
            # Fill data rows
            for row_idx, (display_label, internal) in enumerate(metric_rows, start=2):
                ws.cell(row=row_idx, column=1, value=display_label)
                for ci, rowdata in enumerate(grp_sorted.to_dict('records'), start=2):
                    val = rowdata.get(internal, np.nan)
                    formatted = format_number(display_label, val)
                    ws.cell(row=row_idx, column=ci, value=formatted)
            
            # Set column widths for readability
            ws.column_dimensions['A'].width = 24
            for j in range(2, 2+len(runs_list)):
                ws.column_dimensions[chr(64+j)].width = 14
        
        # Save Excel file with timestamp if needed
        excel_path = Path("messungen_auswertung.xlsx")
        try:
            wb.save(excel_path)
            print(f"✓ Excel report created: {excel_path}")
        except PermissionError:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            alt = Path(f"messungen_auswertung_{ts}.xlsx")
            wb.save(alt)
            print(f"✓ Excel report created: {alt} (original file locked)")
    except Exception as e:
        print(f"Error during Excel export: {e}")

if __name__ == "__main__":
    main()
